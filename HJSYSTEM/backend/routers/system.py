# -*- coding: utf-8 -*-
"""
系统管理路由
"""

from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from datetime import datetime

from backend.database import get_db
from backend.crud import get_component_count, get_logs, create_log_entry
from backend.schemas import SystemStatus, ExcelExportRequest
from backend.excel_handler import export_to_excel
from backend.config import EXPORTS_DIR
from pathlib import Path

router = APIRouter(prefix="/api/system", tags=["system"])


@router.get("/status", response_model=SystemStatus)
async def get_status(db: Session = Depends(get_db)):
    """获取系统状态"""
    from datetime import datetime, timedelta
    from backend.models import User

    # 在线统计：查询最近90秒内有心跳记录的用户数
    ONLINE_TIMEOUT_SECONDS = 90
    timeout_threshold = datetime.now() - timedelta(seconds=ONLINE_TIMEOUT_SECONDS)
    active_count = db.query(User).filter(User.last_login >= timeout_threshold).count()

    return SystemStatus(
        total_components=get_component_count(db),
        online_users=active_count
    )


@router.get("/health")
async def health_check(db: Session = Depends(get_db)):
    """健康检查接口"""
    try:
        from sqlalchemy import text
        db.execute(text("SELECT 1"))
        db_status = "ok"
    except Exception as e:
        db_status = f"error: {str(e)}"

    return {
        "status": "healthy" if db_status == "ok" else "unhealthy",
        "database": db_status,
        "timestamp": datetime.now().isoformat(),
        "version": "2.0.0"
    }


@router.post("/export")
async def export_components(
    export_request: ExcelExportRequest,
    db: Session = Depends(get_db),
    user_ip: str = "",
    user_agent: str = "",
    computer_name: str = ""
):
    """导出元器件数据到Excel"""
    from backend.crud import get_all_components
    from backend.utils import parse_computer_name
    
    components = get_all_components(db)
    
    # 生成文件名
    filename = export_request.filename or f"元器件数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    filepath = EXPORTS_DIR / f"{filename}.xlsx"
    
    # 确保导出目录存在
    EXPORTS_DIR.mkdir(exist_ok=True)
    
    # 导出到Excel
    export_to_excel(components, filepath, title=export_request.title or "元器件报价清单")
    
    # 记录日志
    if not computer_name and user_ip:
        computer_name = parse_computer_name(user_ip, user_agent)
    create_log_entry(db, {
        "action": "export",
        "details": f"导出Excel: {filename}.xlsx, {len(components)} 条记录",
        "user_ip": user_ip,
        "user_agent": user_agent,
        "computer_name": computer_name
    })
    
    from fastapi.responses import FileResponse
    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{filename}.xlsx"
    )


async def export_custom_data(body: dict, db: Session, user_ip: str = "", user_agent: str = "", computer_name: str = ""):
    """导出自定义数据到Excel（用于导出选中项）"""
    import os
    import shutil
    import openpyxl
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
    from backend.utils import parse_computer_name
    
    # 获取数据和参数
    data = body.get('data', [])
    filename = body.get('fileName', f"导出数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    disk_filename = body.get('diskFileName', filename)
    # 使用服务器本地导出目录，避免权限问题
    save_path = str(EXPORTS_DIR)
    title = filename  # 使用文件名作为标题
    
    # 确保保存路径存在
    EXPORTS_DIR.mkdir(exist_ok=True)
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars
    
    # 定义样式
    title_font = Font(name='Times New Roman', size=9, bold=True, color='000000')
    header_font = Font(name='Times New Roman', size=9, bold=False, color='000000')
    data_font = Font(name='Times New Roman', size=9, bold=False, color='000000')
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    header_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(vertical='center')
    
    # 标题行（合并A1:G1）
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    title_cell.border = thin_border
    ws.row_dimensions[1].height = 18
    
    # 表头行
    headers = ['序号', '名称', '型号及规格', '数量', '单价', '小计', '备注']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[2].height = 18
    
    # 数据行
    for row_idx, item in enumerate(data, 3):
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font = data_font
            cell.alignment = header_alignment
        ws.cell(row=row_idx, column=1, value=item.get('序号', row_idx - 2))
        ws.cell(row=row_idx, column=2, value=item.get('名称', ''))
        ws.cell(row=row_idx, column=3, value=item.get('型号及规格', ''))
        ws.cell(row=row_idx, column=4, value=item.get('数量', 1))
        ws.cell(row=row_idx, column=5, value=item.get('单价', 0))
        # 使用公式计算小计
        ws.cell(row=row_idx, column=6, value=f"=D{row_idx}*E{row_idx}")
        ws.cell(row=row_idx, column=7, value=item.get('备注', ''))

        ws.row_dimensions[row_idx].height = 18

    # 添加合计行（紧跟在数据后面，不留空行）
    total_row = len(data) + 3
    total_font = Font(name='Times New Roman', size=9, bold=True, color='000000')
    for col_idx in range(1, 8):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = thin_border
        cell.font = total_font
        cell.alignment = header_alignment
    ws.cell(row=total_row, column=5, value='合计')
    ws.cell(row=total_row, column=6, value=f"=SUM(F3:F{total_row - 1})")
    ws.row_dimensions[total_row].height = 18
    
    # 设置列宽
    FIXED_NARROW_COLS = {'A'}
    MIN_WIDTH_BY_COL = {'B': 6, 'C': 8, 'E': 8, 'F': 8, 'G': 8}
    MAX_WIDTH = 30
    for col_cells in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col_cells:
            if not hasattr(cell, 'column_letter'):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if col_letter in FIXED_NARROW_COLS:
                break
            if cell.value:
                cell_str = str(cell.value)
                cell_len = 0
                for ch in cell_str:
                    if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                        cell_len += 2
                    else:
                        cell_len += 1
                if cell.row == 1 and col_letter and col_letter != 'A':
                    continue
                max_length = max(max_length, cell_len)
        if col_letter:
            if col_letter in FIXED_NARROW_COLS:
                ws.column_dimensions[col_letter].width = 4
            else:
                min_w = MIN_WIDTH_BY_COL.get(col_letter, 6)
                adjusted_width = min(max_length + 1, MAX_WIDTH)
                ws.column_dimensions[col_letter].width = max(adjusted_width, min_w)
    
    # 隐藏H列及之后的列
    for col_letter in 'HIJKLMNOPQRSTUVWXYZ':
        col_dim = ws.column_dimensions.get(col_letter)
        if col_dim is None:
            col_dim = ws.column_dimensions[col_letter]
        col_dim.hidden = True
    
    # 设置打印区域和分页预览
    ws.print_area = f'A1:G{total_row}'
    ws.sheet_view.view = 'pageBreakPreview'
    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 100
    
    # 保存到临时文件
    from fastapi.responses import FileResponse
    final_filepath = os.path.join(save_path, f"{disk_filename}.xlsx")
    wb.save(final_filepath)
    wb.close()

    # 记录日志
    if not computer_name and user_ip:
        computer_name = parse_computer_name(user_ip, user_agent)
    create_log_entry(db, {
        "action": "export",
        "details": f"导出选中Excel: {final_filepath}, {len(data)} 条记录",
        "user_ip": user_ip,
        "user_agent": user_agent,
        "computer_name": computer_name
    })

    # 返回文件流，让浏览器下载
    return FileResponse(
        final_filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{disk_filename}.xlsx"
    )
