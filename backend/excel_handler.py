# -*- coding: utf-8 -*-
"""
HJSYSTEM Excel Handler
Import and export Excel files with openpyxl (.xlsx) and xlrd (.xls)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict, BinaryIO, Tuple
from backend.models import Component
from backend.schemas import ComponentCreate


def _detect_excel_format(file_input) -> str:
    """
    Detect actual Excel format from file header (magic bytes).
    Supports both file path string and file-like objects.
    Returns: 'xlsx', 'xls', or 'unknown'
    """
    try:
        # Check if it's a file path string or file object
        if isinstance(file_input, str):
            with open(file_input, 'rb') as f:
                header = f.read(8)
        else:
            # It's a file-like object
            pos = file_input.tell()
            file_input.seek(0)
            header = file_input.read(8)
            file_input.seek(pos)
        
        # ZIP-based formats (XLSX, XLSM) start with PK (0x504B)
        if header[:2] == b'PK':
            return 'xlsx'
        # OLE2 compound document (XLS) starts with D0CF11E0
        if len(header) >= 8 and header[:4] == b'\xd0\xcf\x11\xe0':
            return 'xls'
        return 'unknown'
    except Exception:
        return 'unknown'


def import_from_excel(file_input) -> Tuple[List[ComponentCreate], int]:
    """
    Import components from Excel file
    Auto-detects format from file header magic bytes.
    Supports both .xlsx (openpyxl) and .xls (xlrd) formats.
    Supports both file path string and file-like objects.
    Returns: (list of ComponentCreate, skipped count)
    """
    # Auto-detect actual format from magic bytes
    detected_format = _detect_excel_format(file_input)
    
    # Check if it's a file path string or file object
    is_file_object = not isinstance(file_input, str)
    
    if detected_format == 'xls':
        try:
            import xlrd
            if is_file_object:
                # xlrd needs a file path, save to temp file
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as temp_file:
                    file_input.seek(0)
                    temp_file.write(file_input.read())
                    temp_path = temp_file.name
                wb = xlrd.open_workbook(temp_path)
                import os
                os.unlink(temp_path)
            else:
                wb = xlrd.open_workbook(file_input)
            ws = wb.sheet_by_index(0)
            is_xlrd = True
            print(f"[Import] Detected .xls (OLE2) format via magic bytes, using xlrd")
        except ImportError:
            raise Exception("xlrd 库未安装，请运行: pip install xlrd")
        except Exception as e:
            raise Exception(f"无法读取 .xls 文件: {str(e)}")
    elif detected_format == 'xlsx':
        # Use openpyxl for .xlsx files
        if is_file_object:
            file_input.seek(0)
        wb = openpyxl.load_workbook(file_input, data_only=True)
        ws = wb.active
        is_xlrd = False
    else:
        raise Exception(f"无法识别的文件格式（非Excel文件），请检查文件是否损坏或为其他格式")
    
    components = []
    skipped = 0
    
    # Helper function to get cell value based on library
    def get_cell_value(row_idx, col_idx):
        if is_xlrd:
            return ws.cell_value(row_idx, col_idx)
        else:
            return ws.cell(row=row_idx + 1, column=col_idx + 1).value
    
    # Helper function to get row values
    def get_row_values(row_idx):
        if is_xlrd:
            return ws.row_values(row_idx)
        else:
            return [cell.value for cell in ws[row_idx + 1]]
    
    # Get number of rows
    def get_max_row():
        if is_xlrd:
            return ws.nrows
        else:
            return ws.max_row
    
    # Find header row
    header_row = None
    for row_idx in range(min(10, get_max_row())):
        row = get_row_values(row_idx)
        if row and any(cell for cell in row if cell):
            header_row = row_idx
            break
    
    if header_row is None:
        return [], 0
    
    # Map column indices
    # Strip all whitespace (including full-width spaces) for robust matching
    headers = [
        str(cell).strip().lower().replace(' ', '').replace('\u3000', '').replace('\t', '')
        if cell is not None and cell else ""
        for cell in get_row_values(header_row)
    ]
    col_map = {}

    for idx, header in enumerate(headers):
        if "序号" in header or "sequence" in header:
            col_map['sequence'] = idx
        elif "名称" in header or "name" in header:
            col_map['name'] = idx
        elif "型号" in header or "model" in header or "规格" in header:
            col_map['model'] = idx
        elif "数量" in header or "quantity" in header:
            col_map['quantity'] = idx
        elif "单价" in header or "price" in header or "unit" in header:
            col_map['unit_price'] = idx
        elif "备注" in header or "remark" in header:
            col_map['remarks'] = idx

    # Fallback defaults for missing columns
    defaults = {
        'sequence': 0,
        'name': 1,
        'model': 2,
        'quantity': 3,
        'unit_price': 4,
        'remarks': 6
    }
    for key, val in defaults.items():
        if key not in col_map:
            col_map[key] = val
    
    # Read data rows
    for row_idx in range(header_row + 1, get_max_row()):
        row = get_row_values(row_idx)
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        
        try:
            name_idx = col_map.get('name', 1)
            name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] else ""
            
            if not name:
                skipped += 1
                continue
            
            # Parse values
            sequence = 0
            if 'sequence' in col_map and col_map['sequence'] < len(row):
                try:
                    sequence = int(row[col_map['sequence']])
                except:
                    pass
            
            model = ""
            if 'model' in col_map and col_map['model'] < len(row):
                model = str(row[col_map['model']]) if row[col_map['model']] else ""
            
            quantity = 1
            if 'quantity' in col_map and col_map['quantity'] < len(row):
                try:
                    quantity = int(float(row[col_map['quantity']]))
                except:
                    pass
            
            unit_price = 0.0
            if 'unit_price' in col_map and col_map['unit_price'] < len(row):
                try:
                    unit_price = float(row[col_map['unit_price']])
                except:
                    pass
            
            remarks = ""
            if 'remarks' in col_map and col_map['remarks'] < len(row):
                remarks = str(row[col_map['remarks']]) if row[col_map['remarks']] else ""
            
            component = ComponentCreate(
                sequence=sequence,
                name=name,
                model=model,
                quantity=quantity,
                unit_price=unit_price,
                remarks=remarks
            )
            components.append(component)
            
        except Exception as e:
            skipped += 1
            continue
    
    # Close workbook if supported
    try:
        wb.close()
    except:
        pass
    return components, skipped


def export_to_excel(
    components: List[Component],
    output_path: str,
    title: str = "元器件报价清单"
) -> str:
    """
    Export components to Excel file
    Returns: output file path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars
    
    # Define styles
    title_font = Font(name='Times New Roman', size=9, bold=True, color='000000')
    header_font = Font(name='Times New Roman', size=9, bold=False, color='000000')
    data_font = Font(name='Times New Roman', size=9, bold=False, color='000000')
    
    title_alignment = Alignment(horizontal='center', vertical='center')
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    header_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Title row (merged A1:G1) - 7 columns total
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    title_cell.border = thin_border  # 添加黑色边框
    ws.row_dimensions[1].height = 18

    # Header row - 7 columns (no ID, no created_at/updated_at)
    headers = ['序号', '名称', '型号及规格', '数量', '单价', '小计', '备注']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[2].height = 18

    # Data rows
    for row_idx, comp in enumerate(components, 3):
        ws.cell(row=row_idx, column=1, value=comp.sequence or row_idx - 2)
        ws.cell(row=row_idx, column=2, value=comp.name)
        ws.cell(row=row_idx, column=3, value=comp.model or "")
        ws.cell(row=row_idx, column=4, value=comp.quantity)
        ws.cell(row=row_idx, column=5, value=comp.unit_price)
        # Use formula for subtotal: quantity * unit_price
        ws.cell(row=row_idx, column=6, value=f"=D{row_idx}*E{row_idx}")
        ws.cell(row=row_idx, column=7, value=comp.remarks or "")

        # Set row height to 18 points
        ws.row_dimensions[row_idx].height = 18

        # Apply styles
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # Add total row
    total_row = len(components) + 3

    # "合计" in column E (单价列)
    total_label_cell = ws.cell(row=total_row, column=5)
    total_label_cell.value = "合计"
    total_label_cell.font = data_font
    total_label_cell.alignment = header_alignment
    total_label_cell.border = thin_border

    # Total formula in column F (小计列)
    total_formula = f"=SUM(F3:F{total_row-1})"
    total_value_cell = ws.cell(row=total_row, column=6)
    total_value_cell.value = total_formula
    total_value_cell.font = data_font
    total_value_cell.alignment = data_alignment
    total_value_cell.border = thin_border

    # Set total row height to 18 points
    ws.row_dimensions[total_row].height = 18

    # Apply borders to remaining cells in total row
    for col_idx in range(1, 8):
        if col_idx not in [5, 6]:  # Skip E and F columns (already styled)
            cell = ws.cell(row=total_row, column=col_idx)
            cell.border = thin_border
    
    # Auto-fit column widths based on content
    # Columns A(序号) and D(数量) fixed at 4 characters
    FIXED_NARROW_COLS = {'A', 'D'}
    # 名称/型号有中文字符，设为最小宽度让表格紧凑
    MIN_WIDTH_BY_COL = {'B': 6, 'C': 8, 'E': 8, 'F': 8, 'G': 8}
    MAX_WIDTH = 30
    for col_cells in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col_cells:
            # Skip MergedCell objects (no column_letter attribute)
            if not hasattr(cell, 'column_letter'):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if col_letter in FIXED_NARROW_COLS:
                break  # Fixed width, no need to scan
            if cell.value:
                cell_str = str(cell.value)
                cell_len = 0
                for ch in cell_str:
                    if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                        cell_len += 2
                    else:
                        cell_len += 1
                # 标题行（合并单元格）只影响A列（已固定），跳过其余列的标题计数
                if cell.row == 1 and col_letter and col_letter != 'A':
                    continue
                max_length = max(max_length, cell_len)
        if col_letter:
            if col_letter in FIXED_NARROW_COLS:
                ws.column_dimensions[col_letter].width = 4
            else:
                min_w = MIN_WIDTH_BY_COL.get(col_letter, 6)
                # 内容宽度 +1 个字符间距，上限 MAX_WIDTH
                adjusted_width = min(max_length + 1, MAX_WIDTH)
                ws.column_dimensions[col_letter].width = max(adjusted_width, min_w)
    
    # 隐藏 H 列及之后的所有列，避免分页预览下显示空白列
    for col_letter in 'HIJKLMNOPQRSTUVWXYZ':
        col_dim = ws.column_dimensions.get(col_letter)
        if col_dim is None:
            col_dim = ws.column_dimensions[col_letter]
        col_dim.hidden = True

    # 设置打印区域为 A-G 列
    total_row = len(components) + 3
    ws.print_area = f'A1:G{total_row}'
    
    # 设置分页预览
    ws.sheet_view.view = 'pageBreakPreview'
    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 100
    
    # Save
    wb.save(output_path)
    wb.close()
    
    return output_path


def migrate_excel_to_db(excel_path: str, db_session) -> Tuple[int, int]:
    """
    Migrate data from Excel to database
    Returns: (imported count, skipped count)
    """
    from backend.crud import bulk_create_components
    
    components, skipped = import_from_excel(excel_path)
    
    if components:
        imported, duplicates = bulk_create_components(db_session, components)
        return imported, skipped + duplicates
    
    return 0, skipped
