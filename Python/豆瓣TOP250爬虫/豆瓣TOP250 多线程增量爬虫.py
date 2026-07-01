import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
import openpyxl
import os
import time
import re
import unicodedata
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

# ===================== 全局配置区 =====================
# 浏览器请求头
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Referer": "https://movie.douban.com/"
}

# 代理池（反爬备用，没有代理可留空，自动走直连）
PROXY_POOL = [
    # "http://127.0.0.1:7890",
]
USE_PROXY = False

# 配色
COLOR_PALETTE = {
    'header_bg': '0F3B5D',
    'line': 'D0D8E8',
    'link_color': '0563C1'
}

# 正则表达式
director_pattern = re.compile(r'导演:\s*(.*?)\s*(?=主演:|$)')
actor_pattern = re.compile(r'主演:\s*(.*?)\s*(?=\d{4}|$)')
year_pattern = re.compile(r'(\d{4})')

# 文件路径
DESKTOP = os.path.expanduser("~/Desktop")
EXCEL_PATH = os.path.join(DESKTOP, "豆瓣TOP250_增量完整榜单.xlsx")
CHART_SAVE_PATH = os.path.join(DESKTOP, "豆瓣数据分析图表.png")

# 全局存储容器
all_movie_data = []
lock = threading.Lock()

# ===================== 工具函数 =====================
def get_proxy():
    """随机获取代理"""
    if not USE_PROXY or len(PROXY_POOL) == 0:
        return None
    import random
    return {"http": random.choice(PROXY_POOL), "https": random.choice(PROXY_POOL)}

def display_width(text):
    """计算字符宽度，适配中英文自动列宽"""
    return sum(2 if unicodedata.east_asian_width(c) in ('F','W') else 1 for c in str(text or ''))

def auto_fit_columns(ws, min_w=8, max_w=50, padding=3):
    """Excel自动适配列宽"""
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        max_width = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            current_width = display_width(cell.value)
            if current_width > max_width:
                max_width = current_width
        final_width = max(min_w, min(max_width * 1.1 + padding, max_w))
        ws.column_dimensions[letter].width = final_width

def get_exist_movie_ids():
    """增量功能：读取已有Excel，获取电影唯一链接用于去重"""
    exist_links = set()
    if not os.path.exists(EXCEL_PATH):
        return exist_links
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 8 and row[7]:
                exist_links.add(row[7])
        wb.close()
    except Exception as e:
        print(f"读取旧表格异常，将新建表格：{e}")
    return exist_links

# ===================== 单页抓取函数（多线程执行） =====================
def crawl_page(page, exist_links):
    start = page * 25
    url = f"https://movie.douban.com/top250?start={start}"
    print(f"【线程{threading.get_ident()}】正在抓取第 {page+1}/10 页")
    proxies = get_proxy()
    try:
        resp = requests.get(url, headers=HEADERS, proxies=proxies, timeout=15)
        resp.raise_for_status()
        resp.encoding = "utf-8"
        soup = BeautifulSoup(resp.text, "lxml")
        items = soup.find_all("div", class_="item")
        page_movie_list = []

        for item in items:
            # 基础信息
            name_tag = item.find("span", class_="title")
            movie_name = name_tag.get_text(strip=True) if name_tag else "无名称"

            score_tag = item.find("span", class_="rating_num")
            movie_score = float(score_tag.get_text(strip=True)) if score_tag else 0.0

            link_tag = item.find("div", class_="hd").find("a")
            detail_link = link_tag["href"] if link_tag and "href" in link_tag.attrs else ""

            # 去重判断
            if detail_link in exist_links:
                print(f"已存在跳过：{movie_name}")
                continue

            info_tag = item.find("div", class_="bd").find("p")
            info_text = info_tag.get_text(strip=True, separator="") if info_tag else ""

            director = director_pattern.search(info_text).group(1) if director_pattern.search(info_text) else "无"
            actor = actor_pattern.search(info_text).group(1) if actor_pattern.search(info_text) else "无"
            release_year = year_pattern.search(info_text).group(1) if year_pattern.search(info_text) else "无"

            area_type_part = info_text.split(release_year)[-1] if release_year != "无" else info_text
            area_type_list = [i.strip() for i in area_type_part.split("/") if i.strip()]
            area = area_type_list[0] if len(area_type_list) >= 1 else "无"
            movie_type = "、".join(area_type_list[1:]) if len(area_type_list) >= 2 else "无"

            desc_tag = item.find("span", class_="inq")
            short_desc = desc_tag.get_text(strip=True) if desc_tag else "无简介"

            # ========== 抓取长影评（新增进阶功能） ==========
            long_review = "暂无热门长影评"
            if detail_link:
                try:
                    review_resp = requests.get(detail_link, headers=HEADERS, proxies=proxies, timeout=8)
                    review_soup = BeautifulSoup(review_resp.text, "lxml")
                    review_tag = review_soup.find("div", class_="review-short")
                    if review_tag:
                        long_review = review_tag.get_text(strip=True)[:300]
                except:
                    pass

            row_data = [
                movie_name, movie_score, release_year,
                director, actor, area, movie_type,
                detail_link, short_desc, long_review
            ]
            page_movie_list.append(row_data)
            print(f"新增入库：{movie_name} | 评分：{movie_score}")

        # 线程锁防止全局列表冲突
        with lock:
            all_movie_data.extend(page_movie_list)
        time.sleep(0.6)
    except Exception as e:
        print(f"第{page+1}页抓取失败：{e}")

# ===================== 写入Excel（增量模式+专业格式） =====================
def save_to_excel(new_data):
    headers = [
        "电影名称", "豆瓣评分", "上映年份",
        "导演", "主演", "地区", "电影类型",
        "电影详情链接", "一句话简介", "热门长影评"
    ]
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "豆瓣TOP250增量榜单"
        ws.append(headers)
        start_row = 2

    # 表头样式
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor=COLOR_PALETTE['header_bg'])
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_gray_side = Side(style='thin', color=COLOR_PALETTE['line'])
    header_border = Border(bottom=thin_gray_side, top=thin_gray_side, left=thin_gray_side, right=thin_gray_side)
    ws.row_dimensions[1].height = 30
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # 写入新增数据+样式
    data_font = Font(name='微软雅黑', size=10, color='333333')
    link_font = Font(name='微软雅黑', color=COLOR_PALETTE['link_color'], underline='single', size=10)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_border = Border(bottom=Side(style='hair', color=COLOR_PALETTE['line']))

    for idx, row_data in enumerate(new_data, start=start_row):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=cell_value)
            cell.font = data_font
            cell.border = data_border
            if col_idx in [1,4,5,7,9,10]:
                cell.alignment = left_align
            elif col_idx == 2:
                cell.alignment = right_align
                cell.number_format = '0.0'
            elif col_idx in [3,6]:
                cell.alignment = center_align
                cell.number_format = '@'
            elif col_idx == 8 and cell_value:
                cell.alignment = left_align
                cell.hyperlink = cell_value
                cell.font = link_font
        ws.row_dimensions[idx].height = 24

    auto_fit_columns(ws)
    ws.freeze_panes = "A2"
    # 打印设置
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.7, bottom=0.7)
    ws.print_title_rows = '1:1'
    wb.save(EXCEL_PATH)
    wb.close()
    print(f"\n✅ 增量写入完成，新增 {len(new_data)} 部电影")
    print(f"✅ 文件路径：{EXCEL_PATH}")

# ===================== 数据分析+绘图函数 =====================
def analysis_and_plot():
    if not os.path.exists(EXCEL_PATH):
        print("暂无数据，无法分析")
        return
    df = pd.read_excel(EXCEL_PATH)
    df = df[df["上映年份"] != "无"]
    df["上映年份"] = pd.to_numeric(df["上映年份"], errors='coerce')

    # 1. 高分占比（>=8.5为高分）
    high_score = df[df["豆瓣评分"] >= 8.5].shape[0]
    total = df.shape[0]
    high_ratio = high_score / total if total > 0 else 0

    # 2. 年份分布统计
    year_count = df["上映年份"].value_counts().sort_index()
    # 3. 地区分布
    area_count = df["地区"].value_counts().head(10)
    # 4. 类型拆分统计
    type_list = []
    for t_str in df["电影类型"].dropna():
        for t in str(t_str).split("、"):
            if t.strip():
                type_list.append(t.strip())
    type_count = pd.Series(type_list).value_counts().head(10)

    # 绘制四张子图
    plt.rcParams["font.sans-serif"] = ["SimHei"]
    plt.rcParams["axes.unicode_minus"] = False
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle("豆瓣TOP250 数据分析总览", fontsize=16, weight="bold")

    # 子图1：高分低分饼图
    axes[0,0].pie([high_score, total-high_score], labels=[f"高分(≥8.5)\n{high_score}部", f"普通分数\n{total-high_score}部"],
                   autopct="%1.1f%%", colors=["#e74c3c","#3498db"], startangle=90)
    axes[0,0].set_title(f"高分电影占比（整体占比{high_ratio:.1%}）")

    # 子图2：年份分布柱状图
    sns.barplot(x=year_count.index.astype(str), y=year_count.values, ax=axes[0,1], color="#2980b9")
    axes[0,1].set_title("各年份电影数量分布")
    axes[0,1].tick_params(axis='x', rotation=45)

    # 子图3：地区TOP10
    sns.barplot(x=area_count.values, y=area_count.index, ax=axes[1,0], color="#27ae60")
    axes[1,0].set_title("电影出品地区TOP10")

    # 子图4：类型TOP10
    sns.barplot(x=type_count.values, y=type_count.index, ax=axes[1,1], color="#8e44ad")
    axes[1,1].set_title("电影类型频次TOP10")

    plt.tight_layout()
    plt.savefig(CHART_SAVE_PATH, dpi=300, bbox_inches="tight")
    plt.close()
    print(f"\n📊 数据分析完成，图表已保存桌面：{CHART_SAVE_PATH}")
    print(f"📊 高分电影(≥8.5)占比：{high_ratio:.2%}")

# ===================== 主入口 =====================
def main():
    print("===== 豆瓣TOP250 多线程增量爬虫启动 =====")
    exist_movie_links = get_exist_movie_ids()
    print(f"检测到旧表格已有 {len(exist_movie_links)} 部电影，自动去重增量抓取")

    # 多线程池（线程数可控，防止封禁，建议4~8）
    thread_num = 6
    with ThreadPoolExecutor(max_workers=thread_num) as executor:
        tasks = [executor.submit(crawl_page, page, exist_movie_links) for page in range(10)]
        for future in as_completed(tasks):
            future.result()

    if len(all_movie_data) > 0:
        save_to_excel(all_movie_data)
    else:
        print("本次没有新增电影，无需写入")

    # 执行数据分析绘图
    analysis_and_plot()
    print("\n🏁 全部任务执行完毕！")

if __name__ == "__main__":
    main()